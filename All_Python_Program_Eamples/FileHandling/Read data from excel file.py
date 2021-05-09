import openpyxl
from openpyxl import Workbook

wb=openpyxl.load_workbook("testExcel.xlsx")
sheet=wb.active #wb.get_sheet_by_name
row=sheet.max_row
col=sheet.max_column
#print(row,col)
print("Eno\t","Ename\t","Esal")
print("-"*30)
for r in range(2,row+1):
    for c in range(1,col+1):
        print(sheet.cell(row=r,column=c).value,"\t",end="")
    print()